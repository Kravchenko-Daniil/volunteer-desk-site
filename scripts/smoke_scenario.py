from io import BytesIO
from pathlib import Path

from django.contrib.auth import get_user_model
from django.test import Client
from django.urls import reverse
from openpyxl import load_workbook

from apps.applications.models import Application, ApplicationAction
from apps.districts.models import District


def expect_status(response, expected, step):
    if response.status_code != expected:
        raise AssertionError(f'{step}: HTTP {response.status_code}, ожидался {expected}')


User = get_user_model()
district, _ = District.objects.get_or_create(
    name='Центральный',
    defaults={'is_active': True},
)
admin, _ = User.objects.get_or_create(
    email='admin@example.test',
    defaults={
        'username': 'smoke-admin',
        'full_name': 'Тестовый администратор',
        'role': User.Role.ADMIN,
    },
)
admin.role = User.Role.ADMIN
admin.is_active = True
admin.set_password('LocalAdmin123!')
admin.save()

volunteer, _ = User.objects.get_or_create(
    email='volunteer@example.test',
    defaults={
        'username': 'smoke-volunteer',
        'full_name': 'Тестовый волонтёр',
        'role': User.Role.VOLUNTEER,
    },
)
volunteer.role = User.Role.VOLUNTEER
volunteer.is_active = True
volunteer.set_password('LocalVolunteer123!')
volunteer.save()

Application.objects.filter(applicant_email__endswith='@smoke.test').delete()

client = Client(HTTP_HOST='localhost', HTTP_USER_AGENT='Codex smoke test')
response = client.post(
    reverse('applications:public_create'),
    {
        'applicant_full_name': 'Тестовый заявитель',
        'applicant_email': 'public@smoke.test',
        'district': district.pk,
        'help_description': ' '.join([
            'Нужна тестовая помощь с доставкой продуктов и сопровождением заявителя.'
        ] * 12),
        'people_needed': 1,
        'comment': 'Сценарная проверка',
        'personal_data_agreed': 'on',
        'honeypot': '',
    },
)
expect_status(response, 302, 'публичная заявка')
application = Application.objects.get(applicant_email='public@smoke.test')
assert application.status == Application.Status.NEW
assert application.actions.filter(action=ApplicationAction.Action.CREATED).exists()

client.force_login(admin)
expect_status(client.get(reverse('applications:list')), 200, 'список администратора')
expect_status(client.get(reverse('applications:detail', args=[application.pk])), 200, 'карточка администратора')
expect_status(client.post(reverse('applications:accept', args=[application.pk])), 302, 'принятие заявки')
application.refresh_from_db()
assert application.status == Application.Status.ACCEPTED

response = client.post(
    reverse('applications:assign', args=[application.pk]),
    {'volunteer_id': volunteer.pk},
)
expect_status(response, 302, 'назначение волонтёра')
application.refresh_from_db()
assert application.status == Application.Status.ASSIGNED
assert application.assigned_volunteer_id == volunteer.pk

client.force_login(volunteer)
expect_status(client.get(reverse('applications:volunteer_list')), 200, 'кабинет волонтёра')
expect_status(client.get(reverse('applications:volunteer_detail', args=[application.pk])), 200, 'карточка волонтёра')
response = client.post(
    reverse('applications:complete', args=[application.pk]),
    {'completion_comment': 'Помощь оказана, сценарий завершён.'},
)
expect_status(response, 302, 'выполнение заявки')
application.refresh_from_db()
assert application.status == Application.Status.COMPLETED
assert application.completed_at is not None

for email, route, extra in [
    ('self@smoke.test', 'applications:volunteer_create_self', {}),
    ('', 'applications:volunteer_create_for_other', {}),
]:
    response = client.post(
        reverse(route),
        {
            'applicant_full_name': 'Заявка волонтёра',
            'applicant_email': email,
            'district': district.pk,
            'help_description': 'Проверка создания заявки волонтёром.',
            'people_needed': 2,
            'comment': '',
            'personal_data_agreed': 'on',
            'honeypot': '',
            **extra,
        },
    )
    expect_status(response, 302, route)
    expected_email = volunteer.email if route == 'applications:volunteer_create_self' else email
    assert Application.objects.filter(applicant_email=expected_email, created_by_user=volunteer).exists()

client.force_login(admin)
artifact_dir = Path('tmp/smoke')
artifact_dir.mkdir(parents=True, exist_ok=True)

pdf_response = client.get(reverse('exports:application_pdf', args=[application.pk]))
expect_status(pdf_response, 200, 'PDF')
assert pdf_response['Content-Type'] == 'application/pdf'
assert pdf_response.content.startswith(b'%PDF-')
pdf_path = artifact_dir / f'application_{application.pk}.pdf'
pdf_path.write_bytes(pdf_response.content)

xlsx_response = client.get(
    reverse('exports:applications_xlsx'),
    {'status': Application.Status.COMPLETED},
)
expect_status(xlsx_response, 200, 'Excel')
xlsx_path = artifact_dir / 'applications_completed.xlsx'
xlsx_path.write_bytes(xlsx_response.content)
workbook = load_workbook(BytesIO(xlsx_response.content), read_only=True, data_only=True)
sheet = workbook['Заявки']
rows = list(sheet.iter_rows(values_only=True))
assert len(rows) == 2, f'Excel-фильтр вернул {len(rows) - 1} строк вместо 1'
assert rows[1][0] == application.pk
assert rows[1][7] == application.get_status_display()

expected_actions = {
    ApplicationAction.Action.CREATED,
    ApplicationAction.Action.ACCEPTED,
    ApplicationAction.Action.ASSIGNED,
    ApplicationAction.Action.COMPLETED,
    ApplicationAction.Action.PDF_GENERATED,
}
actual_actions = set(application.actions.values_list('action', flat=True))
assert expected_actions <= actual_actions

print(f'OK application_id={application.pk}')
print(f'OK district={district.name}')
print(f'OK admin={admin.email}')
print(f'OK volunteer={volunteer.email}')
print(f'OK pdf={pdf_path.resolve()}')
print(f'OK xlsx={xlsx_path.resolve()}')
