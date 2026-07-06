import io

import openpyxl
from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse

from apps.applications.models import Application
from apps.districts.models import District

User = get_user_model()


class ExcelExportTest(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username='admin_test',
            email='admin_test@example.com',
            password='testpass123',
            role=User.Role.ADMIN,
            is_staff=True,
        )

        self.district1 = District.objects.create(name='Район 1')
        self.district2 = District.objects.create(name='Район 2')

        self.app1 = Application.objects.create(
            applicant_full_name='Иванов Иван Иванович',
            applicant_email='ivan@example.com',
            district=self.district1,
            help_description='Нужна помощь с продуктами',
            status=Application.Status.NEW,
        )
        self.app2 = Application.objects.create(
            applicant_full_name='Петров Пётр Петрович',
            applicant_email='petr@example.com',
            district=self.district2,
            help_description='Нужна помощь с лекарствами',
            status=Application.Status.ACCEPTED,
        )
        self.app3 = Application.objects.create(
            applicant_full_name='Сидорова Мария Ивановна',
            applicant_email='maria@example.com',
            district=self.district1,
            help_description='Нужна помощь с уборкой',
            status=Application.Status.REJECTED,
        )

        self.export_url = reverse('exports:applications_xlsx')

    def _load_sheet(self, content):
        wb = openpyxl.load_workbook(io.BytesIO(content))
        return wb['Заявки']

    def test_export_all_no_filters_returns_3_rows(self):
        self.client.force_login(self.admin)
        response = self.client.get(self.export_url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        ws = self._load_sheet(response.content)
        data_rows = ws.max_row - 1  # subtract header row
        self.assertEqual(
            data_rows,
            3,
            msg=f'Expected 3 data rows, got {data_rows}. Bug: export returns 0 rows when no filters.',
        )

    def test_export_filter_by_status_new(self):
        self.client.force_login(self.admin)
        response = self.client.get(self.export_url, {'status': 'new'})
        self.assertEqual(response.status_code, 200)
        ws = self._load_sheet(response.content)
        data_rows = ws.max_row - 1  # subtract header row
        # Only 1 application has status=new
        self.assertEqual(
            data_rows,
            1,
            msg=f'Expected 1 data row for status=new, got {data_rows}.',
        )
        self.assertLessEqual(data_rows, 3)
