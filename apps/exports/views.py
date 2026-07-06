import math

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from weasyprint import HTML
from django.utils import timezone
from apps.applications.models import Application, ApplicationAction
from apps.applications.forms import ApplicationFilterForm
from apps.applications.selectors import filter_applications
from apps.exports.models import ExportLog

User = get_user_model()

def require_admin(user):
    if not user.is_authenticated or not user.is_project_admin:
        raise PermissionDenied

@login_required
def export_applications_xlsx(request):
    require_admin(request.user)
    qs = Application.objects.select_related('district', 'assigned_volunteer')
    form = ApplicationFilterForm(request.GET or None, user_model=User)
    if form.is_valid():
        qs = filter_applications(qs, form.cleaned_data)
    ExportLog.objects.create(created_by_user=request.user, format=ExportLog.Format.XLSX, filters_json=dict(request.GET))

    applications = list(qs)
    ApplicationAction.objects.bulk_create([
        ApplicationAction(
            application=application,
            action=ApplicationAction.Action.EXPORTED,
            changed_by_user=request.user,
            comment='Включена в Excel-экспорт',
        )
        for application in applications
    ])

    wb = Workbook()
    ws = wb.active
    ws.title = 'Заявки'
    ws.append(['№', 'Дата', 'ФИО', 'Email', 'Район', 'Описание помощи', 'Кол-во людей', 'Статус', 'Волонтёр', 'Дата выполнения', 'Комментарий выполнения'])
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='174EA6')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 32
    for app in applications:
        ws.append([
            app.id,
            timezone.localtime(app.created_at).replace(tzinfo=None),
            app.applicant_full_name,
            app.applicant_email,
            app.district.name,
            app.help_description,
            app.people_needed,
            app.get_status_display(),
            str(app.assigned_volunteer or ''),
            timezone.localtime(app.completed_at).replace(tzinfo=None) if app.completed_at else None,
            app.completion_comment,
        ])
        row_number = ws.max_row
        for cell in ws[row_number]:
            cell.alignment = Alignment(vertical='top')
        ws.cell(row_number, 6).alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(row_number, 11).alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(row_number, 2).number_format = 'dd.mm.yyyy hh:mm'
        ws.cell(row_number, 10).number_format = 'dd.mm.yyyy hh:mm'
        wrapped_lines = max(
            math.ceil(len(app.help_description) / 55),
            math.ceil(len(app.completion_comment or '') / 45),
            1,
        )
        ws.row_dimensions[row_number].height = min(390, max(24, wrapped_lines * 15))

    widths = [10, 20, 28, 28, 20, 45, 16, 16, 28, 20, 35]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = '1:1'

    info = wb.create_sheet('Информация')
    info.append(['Дата экспорта', timezone.localtime().replace(tzinfo=None)])
    info.append(['Фильтры', request.GET.urlencode() or 'Без фильтров'])
    info['B1'].number_format = 'dd.mm.yyyy hh:mm'
    for cell in info['A']:
        cell.font = Font(bold=True)
    info.column_dimensions['A'].width = 20
    info.column_dimensions['B'].width = 60
    info.sheet_view.showGridLines = False
    info.page_setup.fitToWidth = 1
    info.page_setup.fitToHeight = 1
    info.sheet_properties.pageSetUpPr.fitToPage = True

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="applications_export.xlsx"'
    wb.save(response)
    messages.success(request, f'Excel сформирован: {len(applications)} заявок.')
    return response

@login_required
def application_pdf(request, pk):
    require_admin(request.user)
    application = get_object_or_404(Application.objects.select_related('district', 'assigned_volunteer'), pk=pk)
    html = render_to_string('applications/application_pdf.html', {'application': application})
    pdf = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
    ApplicationAction.objects.create(application=application, action=ApplicationAction.Action.PDF_GENERATED, changed_by_user=request.user)
    messages.success(request, 'PDF заявки сформирован.')
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="application_{application.pk}.pdf"'
    return response
