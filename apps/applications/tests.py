from io import BytesIO
from unittest import mock

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from apps.districts.models import District

from .models import Application, ApplicationAction
from .models import SubmissionAttempt


User = get_user_model()


@override_settings(RATE_LIMIT_10_MINUTES=100, RATE_LIMIT_24_HOURS=100)
class ProjectTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.active_district = District.objects.create(name='Активный', sort_order=20)
        cls.first_district = District.objects.create(name='Первый', sort_order=10)
        cls.inactive_district = District.objects.create(name='Скрытый', is_active=False, sort_order=1)
        cls.admin = User.objects.create_user(
            username='admin',
            email='admin@test.local',
            password='AdminPassword123!',
            role=User.Role.ADMIN,
            full_name='Администратор',
        )
        cls.volunteer = User.objects.create_user(
            username='volunteer',
            email='volunteer@test.local',
            password='VolunteerPassword123!',
            role=User.Role.VOLUNTEER,
            full_name='Волонтёр',
        )

    def application_data(self, **overrides):
        data = {
            'applicant_full_name': 'Иван Иванов',
            'applicant_email': 'applicant@test.local',
            'district': self.active_district.pk,
            'help_description': 'Нужна помощь с доставкой.',
            'people_needed': 1,
            'comment': '',
            'personal_data_agreed': 'on',
            'honeypot': '',
        }
        data.update(overrides)
        return data

    def create_application(self, **overrides):
        data = {
            'applicant_full_name': 'Иван Иванов',
            'applicant_email': 'applicant@test.local',
            'district': self.active_district,
            'help_description': 'Нужна помощь с доставкой.',
            'personal_data_agreed': True,
        }
        data.update(overrides)
        return Application.objects.create(**data)


class StructureAndPublicFormTests(ProjectTestCase):
    def test_public_pages_are_available(self):
        for route in ['core:home', 'applications:public_create', 'core:privacy', 'core:personal_data_consent']:
            self.assertEqual(self.client.get(reverse(route)).status_code, 200)

    def test_only_active_districts_are_sorted_in_form(self):
        response = self.client.get(reverse('applications:public_create'))
        districts = list(response.context['form'].fields['district'].queryset)
        self.assertEqual(districts, [self.first_district, self.active_district])

    def test_required_fields_and_email_validation(self):
        required = ['applicant_full_name', 'applicant_email', 'district', 'help_description', 'personal_data_agreed']
        for field in required:
            with self.subTest(field=field):
                data = self.application_data()
                data.pop(field)
                response = self.client.post(reverse('applications:public_create'), data)
                self.assertEqual(response.status_code, 200)
                self.assertIn(field, response.context['form'].errors)
        response = self.client.post(
            reverse('applications:public_create'),
            self.application_data(applicant_email='not-an-email'),
        )
        self.assertIn('applicant_email', response.context['form'].errors)

    def test_public_application_saves_audit_data(self):
        response = self.client.post(
            reverse('applications:public_create'),
            self.application_data(),
            REMOTE_ADDR='203.0.113.10',
            HTTP_USER_AGENT='Project test agent',
        )
        self.assertRedirects(response, reverse('applications:thanks'))
        application = Application.objects.get(applicant_email='applicant@test.local')
        self.assertEqual(application.status, Application.Status.NEW)
        self.assertEqual(application.applicant_ip, '203.0.113.10')
        self.assertEqual(application.user_agent, 'Project test agent')
        self.assertIsNotNone(application.personal_data_agreed_at)
        self.assertTrue(application.actions.filter(action=ApplicationAction.Action.CREATED).exists())

    def test_honeypot_silently_discards_submission(self):
        response = self.client.post(
            reverse('applications:public_create'),
            self.application_data(honeypot='spam-link'),
        )
        self.assertRedirects(response, reverse('applications:thanks'))
        self.assertFalse(Application.objects.filter(applicant_email='applicant@test.local').exists())
        self.assertTrue(SubmissionAttempt.objects.filter(was_honeypot=True).exists())

    @override_settings(TRUST_X_FORWARDED_FOR=False)
    def test_forwarded_ip_is_ignored_by_default(self):
        self.client.post(
            reverse('applications:public_create'),
            self.application_data(applicant_email='ip-default@test.local'),
            REMOTE_ADDR='203.0.113.10',
            HTTP_X_FORWARDED_FOR='198.51.100.20',
        )
        self.assertEqual(
            Application.objects.get(applicant_email='ip-default@test.local').applicant_ip,
            '203.0.113.10',
        )

    @override_settings(TRUST_X_FORWARDED_FOR=True)
    def test_forwarded_ip_is_used_only_when_trusted(self):
        self.client.post(
            reverse('applications:public_create'),
            self.application_data(applicant_email='ip-proxy@test.local'),
            REMOTE_ADDR='10.0.0.2',
            HTTP_X_FORWARDED_FOR='198.51.100.20, 10.0.0.1',
        )
        self.assertEqual(
            Application.objects.get(applicant_email='ip-proxy@test.local').applicant_ip,
            '198.51.100.20',
        )

    @override_settings(RATE_LIMIT_10_MINUTES=1, RATE_LIMIT_24_HOURS=20)
    def test_rate_limit_blocks_repeated_submission(self):
        first = self.client.post(
            reverse('applications:public_create'),
            self.application_data(applicant_email='rate-one@test.local'),
            REMOTE_ADDR='203.0.113.44',
        )
        second = self.client.post(
            reverse('applications:public_create'),
            self.application_data(applicant_email='rate-two@test.local'),
            REMOTE_ADDR='203.0.113.44',
        )
        self.assertEqual(first.status_code, 302)
        self.assertEqual(second.status_code, 429)
        self.assertFalse(Application.objects.filter(applicant_email='rate-two@test.local').exists())
        self.assertTrue(SubmissionAttempt.objects.filter(was_blocked=True).exists())

    @override_settings(
        CAPTCHA_ENABLED=True,
        CAPTCHA_PROVIDER='hcaptcha',
        CAPTCHA_SITE_KEY='test-site-key',
        CAPTCHA_SECRET_KEY='test-secret-key',
    )
    @mock.patch('apps.applications.views.verify_captcha', return_value=False)
    def test_enabled_captcha_must_pass(self, verify):
        response = self.client.post(
            reverse('applications:public_create'),
            self.application_data(applicant_email='captcha@test.local'),
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Application.objects.filter(applicant_email='captcha@test.local').exists())
        verify.assert_called_once()

    def test_design_critical_blocks_are_present(self):
        home = self.client.get(reverse('core:home'))
        self.assertContains(home, 'Кому предназначена помощь')
        self.assertContains(home, 'Как это работает')
        public_form = self.client.get(reverse('applications:public_create'))
        self.assertContains(public_form, 'css/style.css')


class RoleAndAccessTests(ProjectTestCase):
    def test_login_redirect_depends_on_role(self):
        for user, destination in [
            (self.admin, 'applications:list'),
            (self.volunteer, 'applications:volunteer_list'),
        ]:
            with self.subTest(role=user.role):
                self.client.force_login(user)
                response = self.client.get(reverse('users:post_login_redirect'))
                self.assertRedirects(response, reverse(destination), fetch_redirect_response=False)
                self.client.logout()

    def test_volunteer_cannot_open_admin_pages(self):
        self.client.force_login(self.volunteer)
        self.assertEqual(self.client.get(reverse('applications:list')).status_code, 403)

    def test_admin_cannot_open_volunteer_cabinet(self):
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get(reverse('applications:volunteer_list')).status_code, 403)

    def test_anonymous_user_is_redirected_to_login(self):
        response = self.client.get(reverse('applications:list'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_volunteer_cannot_use_admin_actions_or_exports(self):
        application = self.create_application()
        self.client.force_login(self.volunteer)
        restricted_requests = [
            self.client.get(reverse('exports:applications_xlsx')),
            self.client.get(reverse('exports:application_pdf', args=[application.pk])),
            self.client.post(reverse('applications:accept', args=[application.pk])),
            self.client.post(reverse('applications:reject', args=[application.pk])),
            self.client.post(
                reverse('applications:assign', args=[application.pk]),
                {'volunteer_id': self.volunteer.pk},
            ),
        ]
        self.assertTrue(all(response.status_code == 403 for response in restricted_requests))

    def test_volunteer_cannot_open_foreign_application(self):
        foreign_volunteer = User.objects.create_user(
            username='foreign',
            email='foreign@test.local',
            password='ForeignPassword123!',
            role=User.Role.VOLUNTEER,
        )
        application = self.create_application(assigned_volunteer=foreign_volunteer)
        self.client.force_login(self.volunteer)
        self.assertEqual(
            self.client.get(reverse('applications:volunteer_detail', args=[application.pk])).status_code,
            403,
        )


class StatusWorkflowTests(ProjectTestCase):
    def setUp(self):
        self.application = self.create_application()
        self.client.force_login(self.admin)

    def test_full_valid_workflow(self):
        self.assertEqual(
            self.client.get(reverse('applications:accept', args=[self.application.pk])).status_code,
            405,
        )
        self.client.post(reverse('applications:accept', args=[self.application.pk]))
        self.application.refresh_from_db()
        self.assertEqual(self.application.status, Application.Status.ACCEPTED)

        self.client.post(
            reverse('applications:assign', args=[self.application.pk]),
            {'volunteer_id': self.volunteer.pk},
        )
        self.application.refresh_from_db()
        self.assertEqual(self.application.status, Application.Status.ASSIGNED)

        self.client.force_login(self.volunteer)
        self.client.post(
            reverse('applications:complete', args=[self.application.pk]),
            {'completion_comment': 'Готово'},
        )
        self.application.refresh_from_db()
        self.assertEqual(self.application.status, Application.Status.COMPLETED)
        self.assertEqual(self.application.completion_comment, 'Готово')
        self.assertEqual(
            list(self.application.actions.values_list('action', flat=True)),
            [
                ApplicationAction.Action.COMPLETED,
                ApplicationAction.Action.ASSIGNED,
                ApplicationAction.Action.ACCEPTED,
            ],
        )

    def test_invalid_transitions_are_rejected(self):
        self.client.post(
            reverse('applications:assign', args=[self.application.pk]),
            {'volunteer_id': self.volunteer.pk},
        )
        self.application.refresh_from_db()
        self.assertEqual(self.application.status, Application.Status.NEW)

        self.client.post(reverse('applications:reject', args=[self.application.pk]))
        self.client.post(
            reverse('applications:assign', args=[self.application.pk]),
            {'volunteer_id': self.volunteer.pk},
        )
        self.application.refresh_from_db()
        self.assertEqual(self.application.status, Application.Status.REJECTED)


class FiltersAndExportTests(ProjectTestCase):
    def setUp(self):
        self.completed = self.create_application(applicant_email='completed@test.local')
        self.completed.transition_to(Application.Status.ACCEPTED)
        self.completed.transition_to(Application.Status.ASSIGNED, assigned_volunteer=self.volunteer)
        self.completed.transition_to(Application.Status.COMPLETED)
        self.create_application(applicant_email='new@test.local')
        self.client.force_login(self.admin)

    def test_admin_list_and_export_use_same_filters(self):
        params = {
            'q': 'completed@',
            'status': Application.Status.COMPLETED,
            'district': self.active_district.pk,
            'volunteer': self.volunteer.pk,
        }
        response = self.client.get(reverse('applications:list'), params)
        self.assertEqual(list(response.context['applications']), [self.completed])
        self.assertContains(response, 'data-testid="adaptive-filters"')
        self.assertContains(response, 'data-testid="mobile-application-cards"')

        response = self.client.get(reverse('exports:applications_xlsx'), params)
        workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
        rows = list(workbook['Заявки'].iter_rows(values_only=True))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][0], self.completed.pk)
        self.assertIn('Дата экспорта', workbook['Информация']['A1'].value)
        self.assertTrue(
            self.completed.actions.filter(
                action=ApplicationAction.Action.EXPORTED,
                changed_by_user=self.admin,
            ).exists()
        )

    def test_excel_handles_empty_result(self):
        response = self.client.get(
            reverse('exports:applications_xlsx'),
            {'q': 'definitely-not-found'},
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
        rows = list(workbook['Заявки'].iter_rows(values_only=True))
        self.assertEqual(len(rows), 1)


class VolunteerApplicationCreationTests(ProjectTestCase):
    def setUp(self):
        self.client.force_login(self.volunteer)

    def test_self_application_uses_profile_identity(self):
        response = self.client.post(
            reverse('applications:volunteer_create_self'),
            self.application_data(
                applicant_full_name='Подменённое имя',
                applicant_email='spoofed@test.local',
            ),
        )
        self.assertRedirects(response, reverse('applications:volunteer_list'))
        application = Application.objects.get(created_from=Application.CreatedFrom.VOLUNTEER_SELF)
        self.assertEqual(application.applicant_full_name, self.volunteer.full_name)
        self.assertEqual(application.applicant_email, self.volunteer.email)
        self.assertEqual(application.created_by_user, self.volunteer)

    def test_other_application_allows_missing_email_and_requires_consent(self):
        data = self.application_data(applicant_email='')
        response = self.client.post(reverse('applications:volunteer_create_for_other'), data)
        self.assertRedirects(response, reverse('applications:volunteer_list'))
        application = Application.objects.get(created_from=Application.CreatedFrom.VOLUNTEER_FOR_OTHER)
        self.assertEqual(application.applicant_email, '')
        self.assertEqual(application.created_by_user, self.volunteer)
        self.assertIsNone(application.assigned_volunteer)

        data = self.application_data(applicant_email='another@test.local')
        data.pop('personal_data_agreed')
        response = self.client.post(reverse('applications:volunteer_create_for_other'), data)
        self.assertEqual(response.status_code, 200)
        self.assertIn('personal_data_agreed', response.context['form'].errors)


class PdfHistoryTests(ProjectTestCase):
    def test_pdf_contains_application_and_records_history(self):
        application = self.create_application(help_description='Длинное описание ' * 100)
        self.client.force_login(self.admin)
        response = self.client.get(reverse('exports:application_pdf', args=[application.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response.content.startswith(b'%PDF-'))
        self.assertTrue(
            application.actions.filter(
                action=ApplicationAction.Action.PDF_GENERATED,
                changed_by_user=self.admin,
            ).exists()
        )
        detail = self.client.get(reverse('applications:detail', args=[application.pk]))
        self.assertContains(detail, 'Распечатать заявку')


class CompletedApplicationSafetyTests(ProjectTestCase):
    def test_completed_application_cannot_be_completed_twice(self):
        application = self.create_application()
        application.transition_to(Application.Status.ACCEPTED)
        application.transition_to(Application.Status.ASSIGNED, assigned_volunteer=self.volunteer)
        self.client.force_login(self.volunteer)
        complete_url = reverse('applications:complete', args=[application.pk])
        self.client.post(complete_url, {'completion_comment': 'Первое выполнение'})
        first_completed_at = Application.objects.get(pk=application.pk).completed_at
        self.client.post(complete_url, {'completion_comment': 'Повторное выполнение'})
        application.refresh_from_db()
        self.assertEqual(application.completed_at, first_completed_at)
        self.assertEqual(application.completion_comment, 'Первое выполнение')
        self.assertEqual(
            application.actions.filter(action=ApplicationAction.Action.COMPLETED).count(),
            1,
        )
